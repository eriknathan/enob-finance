import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

from app_core.constants import MONTH_NAMES_SHORT as MONTH_NAMES

# ── Palette ───────────────────────────────────────────────────────────────────
_C_BANNER_BG  = '000918'   # dark — banner do topo da planilha
_C_HEADER_BG  = '1E293B'   # slate-800 — cabeçalho de colunas
_C_SECTION_BG = '0066FF'   # blue — título de seção
_C_TOTAL_BG   = 'DBEAFE'   # blue-100 — linha de total
_C_ALT_BG     = 'EFF6FF'   # blue-50 — linha par
_C_BORDER     = 'CBD5E1'   # slate-300 — borda padrão
_C_BORDER_HDR = '334155'   # slate-700 — borda no header escuro

_C_GREEN      = '16A34A'
_C_RED        = 'DC2626'
_C_BLUE       = '2563EB'
_C_PURPLE     = '7C3AED'
_C_SLATE      = '475569'

_FONT         = 'Calibri'
_CURRENCY_FMT = '"R$" #,##0.00;[Red]"R$" -#,##0.00'
_DATE_FMT     = 'DD/MM/YYYY'

# Tab colors por tipo de aba
_TAB_ANNUAL   = '0066FF'
_TAB_MONTH    = '059669'
_TAB_CARDS    = '7C3AED'
_TAB_PAY      = 'DC2626'
_TAB_SALE     = '16A34A'
_TAB_LOAN     = 'D97706'


# ── Primitivos de estilo ──────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')


def _thin_border(color=_C_BORDER):
    s = Side(border_style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _bottom_border(color=_C_BORDER):
    return Border(bottom=Side(border_style='thin', color=color))


def _font(color='000000', bold=False, size=10, name=_FONT):
    return Font(name=name, color=color, bold=bold, size=size)


def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── Helpers de célula ─────────────────────────────────────────────────────────

def _banner(ws, row, title, subtitle, n_cols):
    """Faixa de título no topo da planilha."""
    ws.row_dimensions[row].height = 36
    cell = ws.cell(row=row, column=1, value=title)
    cell.fill = _fill(_C_BANNER_BG)
    cell.font = Font(name=_FONT, color='FFFFFF', bold=True, size=14)
    cell.alignment = _align('left', 'center')
    if n_cols > 1:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=n_cols)

    row += 1
    ws.row_dimensions[row].height = 18
    sub = ws.cell(row=row, column=1, value=subtitle)
    sub.fill = _fill('0A1428')
    sub.font = Font(name=_FONT, color='94A3B8', size=9)
    sub.alignment = _align('left', 'center')
    if n_cols > 1:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=n_cols)
    return row + 1


def _section_title(ws, row, title, n_cols):
    """Barra azul de separação de seção."""
    ws.row_dimensions[row].height = 22
    cell = ws.cell(row=row, column=1, value=f'  {title}')
    cell.fill = _fill(_C_SECTION_BG)
    cell.font = Font(name=_FONT, color='FFFFFF', bold=True, size=10)
    cell.alignment = _align('left', 'center')
    if n_cols > 1:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=n_cols)


def _header_row(ws, row, headers):
    ws.row_dimensions[row].height = 26
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = _fill(_C_HEADER_BG)
        cell.font = Font(name=_FONT, color='FFFFFF', bold=True, size=10)
        cell.alignment = _align('center', 'center')
        cell.border = _thin_border(_C_BORDER_HDR)


def _total_row(ws, row, label, n_label_cols, n_total_cols):
    """Linha de total com fundo azul-claro."""
    ws.row_dimensions[row].height = 22
    for col in range(1, n_label_cols + n_total_cols + 1):
        ws.cell(row=row, column=col).fill = _fill(_C_TOTAL_BG)
        ws.cell(row=row, column=col).border = _thin_border('BFDBFE')

    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.font = Font(name=_FONT, bold=True, size=10, color=_C_SLATE)
    label_cell.alignment = _align('left', 'center')
    if n_label_cols > 1:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=n_label_cols)


def _data_row(ws, row, alt=False):
    ws.row_dimensions[row].height = 20
    if alt:
        for col in range(1, 20):
            ws.cell(row=row, column=col).fill = _fill(_C_ALT_BG)


def _money(ws, row, col, value, alt=False, color=None):
    cell = ws.cell(row=row, column=col, value=float(value or 0))
    cell.number_format = _CURRENCY_FMT
    cell.alignment = _align('right', 'center')
    cell.border = _thin_border()
    if alt:
        cell.fill = _fill(_C_ALT_BG)
    if color:
        cell.font = Font(name=_FONT, color=color, size=10)
    else:
        cell.font = Font(name=_FONT, size=10)
    return cell


def _money_total(ws, row, col, value, color=None):
    cell = ws.cell(row=row, column=col, value=float(value or 0))
    cell.number_format = _CURRENCY_FMT
    cell.alignment = _align('right', 'center')
    cell.fill = _fill(_C_TOTAL_BG)
    cell.border = _thin_border('BFDBFE')
    cell.font = Font(name=_FONT, bold=True, size=10,
                     color=color if color else _C_SLATE)
    return cell


def _text(ws, row, col, value, alt=False, align='left'):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = _thin_border()
    cell.font = Font(name=_FONT, size=10)
    cell.alignment = _align(align, 'center')
    if alt:
        cell.fill = _fill(_C_ALT_BG)
    return cell


def _date_cell(ws, row, col, value, alt=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = _DATE_FMT
    cell.alignment = _align('center', 'center')
    cell.border = _thin_border()
    cell.font = Font(name=_FONT, size=10)
    if alt:
        cell.fill = _fill(_C_ALT_BG)
    return cell


def _col_width(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _status_cell(ws, row, col, is_ok, alt=False):
    label = 'Pago' if is_ok else 'Pendente'
    cell = _text(ws, row, col, label, alt, align='center')
    cell.font = Font(name=_FONT, size=10,
                     color=_C_GREEN if is_ok else _C_RED)
    return cell


def _pct_cell(ws, row, col, pct, alt=False):
    cell = ws.cell(row=row, column=col, value=pct)
    cell.number_format = '0.0"%"'
    cell.alignment = _align('center', 'center')
    cell.border = _thin_border()
    cell.font = Font(name=_FONT, size=10,
                     color=_C_GREEN if (pct or 0) >= 100 else _C_RED)
    if alt:
        cell.fill = _fill(_C_ALT_BG)
    return cell


# ── Resumo Anual ──────────────────────────────────────────────────────────────

def _build_annual_summary(wb, user):
    from app_months.models import FinancialMonth

    ws = wb.create_sheet('Resumo Anual')
    ws.sheet_properties.tabColor = _TAB_ANNUAL

    n_cols = 14
    headers = [
        'Ano', 'Mês',
        'Entradas', 'Gastos Var.', 'Desp. Fixas', 'Faturas',
        'Pgtos. Parc.', 'Vendas Parc.', 'Emprést.',
        'Total Saídas', 'Investido', 'Meta', '% Meta', 'Saldo',
    ]
    _col_width(ws, [6, 10, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 10, 14])

    row = _banner(ws, 1, 'EnobFinance — Resumo Anual',
                  'Consolidado de todos os meses registrados', n_cols)

    _header_row(ws, row, headers)
    ws.freeze_panes = f'C{row + 1}'
    ws.auto_filter.ref = f'A{row}:N{row}'
    data_start = row + 1
    row += 1

    months = (
        FinancialMonth.objects
        .filter(user=user)
        .order_by('year', 'month')
    )

    totals = {k: Decimal('0') for k in (
        'entries', 'var', 'fixed', 'invoices',
        'inst_pay', 'inst_sale', 'inst_loan',
        'expenses', 'invested', 'goal',
    )}

    for i, fm in enumerate(months):
        alt = (i % 2 == 1)
        ws.row_dimensions[row].height = 20

        entries   = fm.total_entries()
        var       = fm.total_variable_expenses()
        fixed     = fm.total_fixed_expenses()
        invoices  = fm.total_card_invoices()
        inst_pay  = fm.total_installment_payments()
        inst_sale = fm.total_installment_sales()
        inst_loan = fm.total_installment_loans()
        expenses  = fm.total_expenses()
        invested  = fm.total_investments()
        goal      = fm.investment_goal
        balance   = fm.current_balance()
        pct       = round(float(invested / goal * 100), 1) if goal > 0 else None

        _text(ws, row, 1, fm.year, alt, 'center')
        _text(ws, row, 2, MONTH_NAMES[fm.month], alt, 'center')
        _money(ws, row, 3, entries, alt, _C_GREEN)
        _money(ws, row, 4, var, alt)
        _money(ws, row, 5, fixed, alt)
        _money(ws, row, 6, invoices, alt, _C_PURPLE)
        _money(ws, row, 7, inst_pay, alt)
        _money(ws, row, 8, inst_sale, alt, _C_GREEN)
        _money(ws, row, 9, inst_loan, alt)
        _money(ws, row, 10, expenses, alt, _C_RED)
        _money(ws, row, 11, invested, alt, _C_BLUE)
        _money(ws, row, 12, goal, alt)
        _pct_cell(ws, row, 13, pct, alt)

        bal_cell = _money(ws, row, 14, balance, alt,
                          _C_GREEN if balance >= 0 else _C_RED)
        bal_cell.font = Font(name=_FONT, bold=True, size=10,
                             color=_C_GREEN if balance >= 0 else _C_RED)

        totals['entries']   += entries
        totals['var']       += var
        totals['fixed']     += fixed
        totals['invoices']  += invoices
        totals['inst_pay']  += inst_pay
        totals['inst_sale'] += inst_sale
        totals['inst_loan'] += inst_loan
        totals['expenses']  += expenses
        totals['invested']  += invested
        totals['goal']      += goal
        row += 1

    ws.auto_filter.ref = f'A{data_start - 1}:N{row - 1}'

    # Linha de totais
    _total_row(ws, row, 'TOTAL GERAL', 2, 0)
    _money_total(ws, row, 3,  totals['entries'],   _C_GREEN)
    _money_total(ws, row, 4,  totals['var'])
    _money_total(ws, row, 5,  totals['fixed'])
    _money_total(ws, row, 6,  totals['invoices'],  _C_PURPLE)
    _money_total(ws, row, 7,  totals['inst_pay'])
    _money_total(ws, row, 8,  totals['inst_sale'], _C_GREEN)
    _money_total(ws, row, 9,  totals['inst_loan'])
    _money_total(ws, row, 10, totals['expenses'],  _C_RED)
    _money_total(ws, row, 11, totals['invested'],  _C_BLUE)
    _money_total(ws, row, 12, totals['goal'])


# ── Planilha mensal ───────────────────────────────────────────────────────────

def _build_month_sheet(wb, fm):
    month_name = MONTH_NAMES[fm.month]
    sheet_name = f'{month_name}-{str(fm.year)[2:]}'
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = _TAB_MONTH
    _col_width(ws, [32, 14, 14, 16])

    row = _banner(ws, 1,
                  f'{month_name} {fm.year}',
                  f'Lançamentos do mês · Saldo: R$ {float(fm.current_balance()):,.2f}',
                  4)

    # ── Entradas
    _section_title(ws, row, 'Entradas', 3)
    row += 1
    _header_row(ws, row, ['Descrição', 'Data', 'Valor'])
    row += 1
    for i, entry in enumerate(fm.entries.all()):
        alt = i % 2 == 1
        _text(ws, row, 1, entry.description, alt)
        _date_cell(ws, row, 2, entry.date, alt)
        _money(ws, row, 3, entry.amount, alt, _C_GREEN)
        row += 1
    _total_row(ws, row, 'Total Entradas', 2, 0)
    _money_total(ws, row, 3, fm.total_entries(), _C_GREEN)
    row += 2

    # ── Gastos Variáveis
    _section_title(ws, row, 'Gastos Variáveis', 3)
    row += 1
    _header_row(ws, row, ['Descrição', 'Data', 'Valor'])
    row += 1
    for i, exp in enumerate(fm.variable_expenses.all()):
        alt = i % 2 == 1
        _text(ws, row, 1, exp.description, alt)
        _date_cell(ws, row, 2, exp.date, alt)
        _money(ws, row, 3, exp.amount, alt)
        row += 1
    _total_row(ws, row, 'Total Gastos Variáveis', 2, 0)
    _money_total(ws, row, 3, fm.total_variable_expenses(), _C_RED)
    row += 2

    # ── Despesas Fixas
    _section_title(ws, row, 'Despesas Fixas', 3)
    row += 1
    _header_row(ws, row, ['Descrição', 'Valor', 'Status'])
    row += 1
    for i, exp in enumerate(fm.fixed_expenses.all()):
        alt = i % 2 == 1
        _text(ws, row, 1, exp.description, alt)
        _money(ws, row, 2, exp.amount, alt)
        _status_cell(ws, row, 3, exp.is_paid, alt)
        row += 1
    _total_row(ws, row, 'Total Despesas Fixas', 1, 0)
    _money_total(ws, row, 2, fm.total_fixed_expenses(), _C_RED)
    _text(ws, row, 3, '', False)
    row += 2

    # ── Faturas de Cartão
    _section_title(ws, row, 'Faturas de Cartão', 4)
    row += 1
    _header_row(ws, row, ['Cartão', 'Bandeira', 'Valor', 'Status'])
    row += 1
    for i, inv in enumerate(fm.card_invoices.select_related('card').all()):
        alt = i % 2 == 1
        _text(ws, row, 1, inv.card.name, alt)
        _text(ws, row, 2, inv.card.brand, alt)
        _money(ws, row, 3, inv.amount, alt, _C_PURPLE)
        _status_cell(ws, row, 4, inv.status == 'paid', alt)
        row += 1
    _total_row(ws, row, 'Total Faturas', 2, 0)
    _money_total(ws, row, 3, fm.total_card_invoices(), _C_PURPLE)
    _text(ws, row, 4, '', False)
    row += 2

    # ── Investimentos
    _section_title(ws, row, 'Investimentos', 3)
    row += 1
    _header_row(ws, row, ['Onde', 'Data', 'Valor'])
    row += 1
    for i, inv in enumerate(fm.investments.all()):
        alt = i % 2 == 1
        _text(ws, row, 1, inv.place, alt)
        _date_cell(ws, row, 2, inv.date, alt)
        _money(ws, row, 3, inv.amount, alt, _C_BLUE)
        row += 1
    _total_row(ws, row, 'Total Investido', 2, 0)
    _money_total(ws, row, 3, fm.total_investments(), _C_BLUE)
    if fm.investment_goal:
        row += 1
        ws.row_dimensions[row].height = 20
        _text(ws, row, 1, 'Meta de Investimento').font = Font(
            name=_FONT, size=10, color=_C_SLATE)
        _money(ws, row, 3, fm.investment_goal, color=_C_SLATE)
    row += 2

    # ── Parcelamentos do mês
    installments = fm.installments.select_related('plan').all()
    if installments.exists():
        _section_title(ws, row, 'Parcelamentos do Mês', 4)
        row += 1
        _header_row(ws, row, ['Plano', 'Tipo', 'Parcela', 'Valor'])
        row += 1
        for i, inst in enumerate(installments):
            alt = i % 2 == 1
            _text(ws, row, 1, inst.plan.name, alt)
            _text(ws, row, 2, inst.plan.get_kind_display(), alt)
            _text(ws, row, 3, f'{inst.number}/{inst.plan.count_total()}', alt, 'center')
            _money(ws, row, 4, inst.amount, alt)
            row += 1
        row += 1

    # ── Resumo do Mês
    _section_title(ws, row, 'Resumo do Mês', 4)
    row += 1
    summary = [
        ('Saldo anterior',    fm.previous_balance(),   _C_SLATE),
        ('Total de entradas', fm.total_entries(),       _C_GREEN),
        ('Total de saídas',   fm.total_expenses(),      _C_RED),
        ('Total investido',   fm.total_investments(),   _C_BLUE),
        ('Saldo final',       fm.current_balance(),     None),
    ]
    for label, value, color in summary:
        ws.row_dimensions[row].height = 22
        is_final = label == 'Saldo final'
        bal_color = (_C_GREEN if value >= 0 else _C_RED) if is_final else color

        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(name=_FONT, bold=is_final, size=10, color=_C_SLATE)
        label_cell.border = _thin_border()
        label_cell.alignment = _align('left', 'center')
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=3)

        val_cell = _money(ws, row, 4, value, color=bal_color)
        val_cell.font = Font(name=_FONT, bold=is_final, size=10, color=bal_color)
        row += 1


# ── Cartões ───────────────────────────────────────────────────────────────────

def _build_cards_sheet(wb, user):
    from app_cards.models import Card

    ws = wb.create_sheet('Cartões')
    ws.sheet_properties.tabColor = _TAB_CARDS
    _col_width(ws, [32, 22, 16, 16])

    row = _banner(ws, 1, 'Cartões de Crédito',
                  'Cadastro de cartões e dias de fechamento/vencimento', 4)

    _header_row(ws, row, ['Nome', 'Bandeira', 'Fechamento', 'Vencimento'])
    row += 1

    cards = list(Card.objects.filter(user=user).order_by('name'))
    for i, card in enumerate(cards):
        alt = i % 2 == 1
        ws.row_dimensions[row].height = 20
        _text(ws, row, 1, card.name, alt)
        _text(ws, row, 2, card.brand, alt)
        _text(ws, row, 3, f'Dia {card.closing_day}', alt, 'center')
        _text(ws, row, 4, f'Dia {card.due_day}', alt, 'center')
        row += 1

    if cards:
        ws.auto_filter.ref = f'A{row - len(cards) - 1}:D{row - 1}'


# ── Parcelamentos ─────────────────────────────────────────────────────────────

def _build_installments_sheet(wb, user, kind, sheet_name, tab_color):
    from app_installments.models import InstallmentPlan

    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = tab_color
    _col_width(ws, [32, 14, 14, 14, 14, 16])

    subtitle_map = {
        'payment': 'Pagamentos parcelados realizados',
        'sale':    'Vendas parceladas realizadas',
        'loan':    'Empréstimos e financiamentos',
    }
    row = _banner(ws, 1, sheet_name,
                  subtitle_map.get(kind, ''), 6)

    plans = (
        InstallmentPlan.objects
        .filter(user=user, kind=kind)
        .prefetch_related('installments__financial_month')
        .order_by('name')
    )

    for plan in plans:
        # Cabeçalho do plano
        _section_title(ws, row, plan.name, 6)
        row += 1

        # Linha de resumo do plano
        ws.row_dimensions[row].height = 22
        for col, (label, value, color) in enumerate([
            ('Total',    plan.total(),           None),
            ('Pago',     plan.total_paid(),      _C_GREEN),
            ('Restante', plan.total_remaining(), _C_RED),
        ], start=1):
            lc = ws.cell(row=row, column=(col - 1) * 2 + 1, value=label)
            lc.font = Font(name=_FONT, bold=True, size=9, color=_C_SLATE)
            lc.alignment = _align('center', 'center')
            lc.fill = _fill('F1F5F9')
            lc.border = _thin_border()

            vc = ws.cell(row=row, column=(col - 1) * 2 + 2,
                         value=float(plan.total() if label == 'Total'
                                     else plan.total_paid() if label == 'Pago'
                                     else plan.total_remaining()))
            vc.number_format = _CURRENCY_FMT
            vc.font = Font(name=_FONT, bold=True, size=10,
                           color=color if color else _C_SLATE)
            vc.alignment = _align('right', 'center')
            vc.fill = _fill('F1F5F9')
            vc.border = _thin_border()

        parcelas_cell = ws.cell(row=row, column=5,
                                value=f'Parcelas: {plan.count_paid()}/{plan.count_total()}')
        parcelas_cell.font = Font(name=_FONT, bold=True, size=10, color=_C_SLATE)
        parcelas_cell.alignment = _align('center', 'center')
        parcelas_cell.fill = _fill('F1F5F9')
        parcelas_cell.border = _thin_border()
        ws.merge_cells(start_row=row, start_column=5,
                       end_row=row, end_column=6)
        row += 1

        # Tabela de parcelas
        _header_row(ws, row, ['#', 'Mês', 'Ano', 'Valor', 'Status', 'Comprovante'])
        row += 1
        for i, inst in enumerate(plan.installments.all()):
            alt = i % 2 == 1
            ws.row_dimensions[row].height = 20
            _text(ws, row, 1, inst.number, alt, 'center')
            _text(ws, row, 2, MONTH_NAMES[inst.financial_month.month], alt, 'center')
            _text(ws, row, 3, inst.financial_month.year, alt, 'center')
            _money(ws, row, 4, inst.amount, alt)
            _status_cell(ws, row, 5, inst.is_paid, alt)
            receipt = ws.cell(row=row, column=6, value=inst.receipt_url or '—')
            receipt.font = Font(name=_FONT, size=9, color=_C_SLATE)
            receipt.alignment = _align('left', 'center')
            receipt.border = _thin_border()
            if alt:
                receipt.fill = _fill(_C_ALT_BG)
            row += 1

        row += 1


# ── Entry point ───────────────────────────────────────────────────────────────

def build_export(user):
    wb = Workbook()
    wb.remove(wb.active)

    _build_annual_summary(wb, user)

    from app_months.models import FinancialMonth
    months = (
        FinancialMonth.objects
        .filter(user=user)
        .order_by('year', 'month')
    )
    for fm in months:
        _build_month_sheet(wb, fm)

    _build_cards_sheet(wb, user)
    _build_installments_sheet(wb, user, 'payment', 'Pagamentos', _TAB_PAY)
    _build_installments_sheet(wb, user, 'sale',    'Vendas',     _TAB_SALE)
    _build_installments_sheet(wb, user, 'loan',    'Empréstimos', _TAB_LOAN)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
