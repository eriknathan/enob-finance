import io
import zipfile
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from app_cards.models import Card, CardInvoice
from app_installments.models import Installment, InstallmentPlan
from app_months.models import Entry, FinancialMonth, FixedExpense, VariableExpense

from .services import build_export

User = get_user_model()


def make_user(email='test@test.com', password='pass1234'):
    return User.objects.create_user(email=email, password=password)


class BuildExportTests(TestCase):
    def setUp(self):
        self.user = make_user()
        self.fm = FinancialMonth.objects.create(user=self.user, year=2026, month=6)
        Entry.objects.create(financial_month=self.fm, description='Salário', amount=Decimal('5000'))
        VariableExpense.objects.create(financial_month=self.fm, description='Mercado', amount=Decimal('400'))
        FixedExpense.objects.create(
            financial_month=self.fm, description='Aluguel', amount=Decimal('1500'), status='paid'
        )
        self.card = Card.objects.create(
            user=self.user, name='Nubank', brand='Visa', closing_day=10, due_day=20
        )
        CardInvoice.objects.create(card=self.card, financial_month=self.fm, amount=Decimal('300'))

        plan = InstallmentPlan.objects.create(
            user=self.user, name='TV', kind=InstallmentPlan.KIND_PAYMENT
        )
        Installment.objects.create(
            plan=plan, financial_month=self.fm, number=1,
            amount=Decimal('200'), status=Installment.STATUS_UNPAID,
        )
        sale_plan = InstallmentPlan.objects.create(
            user=self.user, name='Carro', kind=InstallmentPlan.KIND_SALE
        )
        Installment.objects.create(
            plan=sale_plan, financial_month=self.fm, number=1,
            amount=Decimal('1000'), status=Installment.STATUS_PAID,
        )

    def test_returns_nonempty_bytes(self):
        buf = build_export(self.user)
        self.assertIsInstance(buf, io.BytesIO)
        data = buf.read()
        self.assertGreater(len(data), 0)

    def test_output_is_valid_xlsx(self):
        buf = build_export(self.user)
        data = buf.read()
        zf = zipfile.ZipFile(io.BytesIO(data))
        self.assertIn('xl/workbook.xml', zf.namelist())

    def test_sheet_names(self):
        buf = build_export(self.user)
        wb = load_workbook(buf)
        names = wb.sheetnames
        self.assertIn('Resumo Anual', names)
        self.assertIn('Cartões', names)
        self.assertIn('Pagamentos', names)
        self.assertIn('Vendas', names)

    def test_month_sheet_created(self):
        buf = build_export(self.user)
        wb = load_workbook(buf)
        # Jun 2026 → sheet name Jun-26
        self.assertIn('Jun-26', wb.sheetnames)

    def test_annual_summary_has_data_row(self):
        buf = build_export(self.user)
        wb = load_workbook(buf)
        ws = wb['Resumo Anual']
        # Row 1 is headers, row 2 should be the June data
        self.assertEqual(ws.cell(row=2, column=1).value, 2026)  # year
        self.assertEqual(ws.cell(row=2, column=2).value, 'Jun')  # month

    def test_export_isolated_per_user(self):
        other = make_user('other@test.com', 'pass1234')
        other_fm = FinancialMonth.objects.create(user=other, year=2026, month=6)
        Entry.objects.create(
            financial_month=other_fm, description='Other income', amount=Decimal('99999')
        )
        buf = build_export(self.user)
        wb = load_workbook(buf)
        ws = wb['Jun-26']
        # Read all cell values from the sheet
        all_values = [ws.cell(row=r, column=c).value
                      for r in range(1, ws.max_row + 1)
                      for c in range(1, ws.max_column + 1)]
        self.assertNotIn('Other income', all_values)
        self.assertIn('Salário', all_values)


class ExportViewTests(TestCase):
    def setUp(self):
        self.user = make_user()

    def test_export_requires_login(self):
        response = self.client.get(reverse('export-excel'))
        self.assertEqual(response.status_code, 302)

    def test_export_returns_xlsx_response(self):
        self.client.login(username='test@test.com', password='pass1234')
        response = self.client.get(reverse('export-excel'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('attachment', response['Content-Disposition'])
        self.assertIn('.xlsx', response['Content-Disposition'])
